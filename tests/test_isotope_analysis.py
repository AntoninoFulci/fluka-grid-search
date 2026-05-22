import pytest
import pandas as pd
from grid_search.isotope_analysis import (
    isotope_symbol,
    molar_mass,
    half_life,
    format_decay_time,
    _build_pivot_sheet,
)
from openpyxl import load_workbook
from grid_search.config import IsotopeConfig
from grid_search.isotope_analysis import run_isotope_analysis


def test_isotope_symbol_co60():
    assert isotope_symbol(27, 60) == "Co-60"


def test_isotope_symbol_cs137():
    assert isotope_symbol(55, 137) == "Cs-137"


def test_molar_mass_co60_positive():
    mass = molar_mass(27, 60)
    assert isinstance(mass, float)
    assert mass > 50.0


def test_molar_mass_unknown_returns_zero():
    assert molar_mass(999, 999) == 0.0


def test_half_life_co60_positive():
    hl = half_life(27, 60)
    assert isinstance(hl, float)
    assert hl > 0.0


def test_half_life_unknown_returns_zero():
    assert half_life(999, 999) == 0.0


def test_half_life_stable_isotope_returns_zero():
    # Fe-56 is stable; radioactivedecay returns "stable" for its half_life()
    assert half_life(26, 56) == 0.0


def test_format_decay_time_seconds():
    assert format_decay_time(30) == "30 s"


def test_format_decay_time_minutes():
    assert format_decay_time(120) == "2 min"


def test_format_decay_time_hours():
    assert format_decay_time(7200) == "2 h"


def test_format_decay_time_days():
    assert format_decay_time(3 * 86400) == "3.0 d"


def test_format_decay_time_zero():
    assert format_decay_time(0) == "0 s"


import struct
from unittest.mock import patch, MagicMock
from grid_search.isotope_analysis import read_resnuclei_file
from grid_search.resnuclei import Detector
from grid_search.state import StateManager


def make_float_bytes(*values):
    return struct.pack(f"={len(values)}f", *values)


def test_missing_rnc_file_is_skipped(tmp_path):
    row = read_resnuclei_file(
        path=tmp_path / "nonexistent",
        requested_isotopes={27: 60},
        params={},
    )
    assert row is None


def test_read_resnuclei_file_returns_known_isotope(tmp_path):
    # Detector: zhigh=2, mhigh=2, nmzmin=0, volume=2.0
    # For Z=1 (z=0), A=3: m = 3 - 2*0 - 0 - 3 = 0, pos = 0 + 0*2 = 0
    # Bq = fdata[0] * volume = 100.0 * 2.0 = 200.0
    # BqErr = edata[0] * fdata[0] * volume = 0.05 * 100.0 * 2.0 = 10.0
    # % Error = (10.0 / 200.0) * 100 = 5.0

    det = Detector(num=1, name="test", volume=2.0, mhigh=2, zhigh=2, nmzmin=0)
    fdata_bytes = make_float_bytes(100.0, 50.0, 30.0, 20.0)
    edata_bytes = make_float_bytes(0.05, 0.10, 0.15, 0.20)

    mock_resn = MagicMock()
    mock_resn.detector = [det]
    mock_resn.tdecay = 86400.0  # 1 day in seconds
    mock_resn.read_data.return_value = fdata_bytes
    mock_resn.read_stat.return_value = (None, None, None, None, None, edata_bytes, None)

    rnc_path = tmp_path / "merged_21"
    rnc_path.write_bytes(b"")  # must exist for path.exists() check

    with patch("grid_search.isotope_analysis.Resnuclei", return_value=mock_resn):
        row = read_resnuclei_file(
            path=rnc_path,
            requested_isotopes={1: 3},  # H-3 (tritium): Z=1, A=3
            params={"beame": 0.05},
        )

    assert row is not None
    assert row["H-3 (Bq)"] == pytest.approx(200.0)
    assert row["H-3 (% Error)"] == pytest.approx(5.0)
    assert "d" in row["CoolingTime"]
    assert "beame=0.05" in row["Parameters"]


def test_read_resnuclei_file_isotope_not_present_gives_zero(tmp_path):
    det = Detector(num=1, name="test", volume=1.0, mhigh=2, zhigh=2, nmzmin=0)
    fdata_bytes = make_float_bytes(0.0, 0.0, 0.0, 0.0)
    edata_bytes = make_float_bytes(0.0, 0.0, 0.0, 0.0)

    mock_resn = MagicMock()
    mock_resn.detector = [det]
    mock_resn.tdecay = 0.0
    mock_resn.read_data.return_value = fdata_bytes
    mock_resn.read_stat.return_value = (None, None, None, None, None, edata_bytes, None)

    rnc_path = tmp_path / "merged_21"
    rnc_path.write_bytes(b"")

    with patch("grid_search.isotope_analysis.Resnuclei", return_value=mock_resn):
        row = read_resnuclei_file(
            path=rnc_path,
            requested_isotopes={27: 60},  # Co-60 not present in this small detector
            params={},
        )

    assert row is not None
    assert row["Co-60 (Bq)"] == pytest.approx(0.0)
    assert row["Co-60 (% Error)"] == pytest.approx(0.0)


def test_run_isotope_analysis_writes_excel(tmp_path):
    ia = IsotopeConfig(isotopes={27: 60}, rnc_files=["merged_21"], output="isotopes.xlsx")
    config = MagicMock()
    config.isotope_analysis = ia

    state = StateManager(tmp_path / "state.json")
    state.data = {
        "beame0.05_matGALLIUM": {
            "parameters": {"beame": 0.05, "mat": "GALLIUM"},
            "runs": {},
        }
    }

    fake_row = {
        "_tdecay_s": 86400.0,
        "CoolingTime": "1.0 d",
        "Parameters": "beame=0.05 mat=GALLIUM",
        "Co-60 (Bq)": 1000.0,
        "Co-60 (% Error)": 5.0,
        "Co-60 (µg)": 0.42,
    }

    with patch("grid_search.isotope_analysis.read_resnuclei_file", return_value=fake_row):
        run_isotope_analysis(tmp_path, config, state)

    output = tmp_path / "isotopes.xlsx"
    assert output.exists()
    df = pd.read_excel(output, sheet_name="beame0.05_matGALLIUM")
    assert "CoolingTime" in df.columns
    assert "Co-60 (Bq)" in df.columns
    assert df["Co-60 (Bq)"].iloc[0] == pytest.approx(1000.0)


def test_run_isotope_analysis_no_files_prints_warning(tmp_path, capsys):
    ia = IsotopeConfig(isotopes={27: 60}, rnc_files=["merged_21"])
    config = MagicMock()
    config.isotope_analysis = ia

    state = StateManager(tmp_path / "state.json")
    state.data = {"beame0.05_matGALLIUM": {"parameters": {}, "runs": {}}}

    with patch("grid_search.isotope_analysis.read_resnuclei_file", return_value=None):
        run_isotope_analysis(tmp_path, config, state)

    assert not (tmp_path / "isotopes.xlsx").exists()
    out = capsys.readouterr().out
    assert "No data found" in out


def test_run_isotope_analysis_combo_filter(tmp_path):
    ia = IsotopeConfig(isotopes={27: 60}, rnc_files=["merged_21"])
    config = MagicMock()
    config.isotope_analysis = ia

    state = StateManager(tmp_path / "state.json")
    state.data = {
        "beame0.05_matGALLIUM": {"parameters": {}, "runs": {}},
        "beame0.1_matGALLIUM": {"parameters": {}, "runs": {}},
    }

    calls = []

    def fake_read(path, isotopes, params):
        calls.append(str(path))
        return None

    with patch("grid_search.isotope_analysis.read_resnuclei_file", side_effect=fake_read):
        run_isotope_analysis(tmp_path, config, state, combo="beame0.05_matGALLIUM")

    assert len(calls) == 1
    assert "beame0.05_matGALLIUM" in calls[0]


def test_tdecay_s_not_in_combo_sheets(tmp_path):
    ia = IsotopeConfig(isotopes={27: 60}, rnc_files=["merged_21"])
    config = MagicMock()
    config.isotope_analysis = ia

    state = StateManager(tmp_path / "state.json")
    state.data = {"beame0.05_matGALLIUM": {"parameters": {}, "runs": {}}}

    fake_row = {
        "_tdecay_s": 86400.0,
        "CoolingTime": "1.0 d",
        "Parameters": "",
        "Co-60 (Bq)": 1000.0,
        "Co-60 (% Error)": 5.0,
        "Co-60 (µg)": 0.42,
    }

    with patch("grid_search.isotope_analysis.read_resnuclei_file", return_value=fake_row):
        run_isotope_analysis(tmp_path, config, state)

    df = pd.read_excel(tmp_path / "isotopes.xlsx", sheet_name="beame0.05_matGALLIUM")
    assert "_tdecay_s" not in df.columns


def test_summary_sheet_is_last_sheet(tmp_path):
    ia = IsotopeConfig(isotopes={27: 60}, rnc_files=["merged_21"], volume=5.0)
    config = MagicMock()
    config.isotope_analysis = ia

    state = StateManager(tmp_path / "state.json")
    state.data = {"beame0.05_matGALLIUM": {"parameters": {"beame": 0.05}, "runs": {}}}

    fake_row = {
        "_tdecay_s": 86400.0,
        "CoolingTime": "1.0 d",
        "Parameters": "beame=0.05",
        "Co-60 (Bq)": 1000.0,
        "Co-60 (% Error)": 5.0,
        "Co-60 (µg)": 0.42,
    }

    with patch("grid_search.isotope_analysis.read_resnuclei_file", return_value=fake_row):
        run_isotope_analysis(tmp_path, config, state)

    xl = pd.ExcelFile(tmp_path / "isotopes.xlsx")
    assert xl.sheet_names[-1] == "Summary"


def test_summary_sheet_bq_values(tmp_path):
    ia = IsotopeConfig(isotopes={27: 60}, rnc_files=["merged_21"], volume=5.0)
    config = MagicMock()
    config.isotope_analysis = ia

    state = StateManager(tmp_path / "state.json")
    state.data = {"beame0.05_matGALLIUM": {"parameters": {"beame": 0.05}, "runs": {}}}

    fake_row = {
        "_tdecay_s": 86400.0,
        "CoolingTime": "1.0 d",
        "Parameters": "beame=0.05",
        "Co-60 (Bq)": 1000.0,
        "Co-60 (% Error)": 5.0,
        "Co-60 (µg)": 0.42,
    }

    with patch("grid_search.isotope_analysis.read_resnuclei_file", return_value=fake_row):
        run_isotope_analysis(tmp_path, config, state)

    # Activity (Bq) title at sheet row 1; DataFrame header at sheet row 2 = pandas header index 1
    df = pd.read_excel(tmp_path / "isotopes.xlsx", sheet_name="Summary", header=1)
    assert "Co-60 (Bq)" in df.columns
    assert df["Co-60 (Bq)"].iloc[0] == pytest.approx(1000.0)
    assert "beame" in df.columns
    assert df["beame"].iloc[0] == pytest.approx(0.05)


def test_summary_sheet_normalized_values(tmp_path):
    ia = IsotopeConfig(isotopes={27: 60}, rnc_files=["merged_21"], volume=5.0)
    config = MagicMock()
    config.isotope_analysis = ia

    state = StateManager(tmp_path / "state.json")
    state.data = {"beame0.05_matGALLIUM": {"parameters": {"beame": 0.05}, "runs": {}}}

    fake_row = {
        "_tdecay_s": 86400.0,
        "CoolingTime": "1.0 d",
        "Parameters": "beame=0.05",
        "Co-60 (Bq)": 1000.0,
        "Co-60 (% Error)": 5.0,
        "Co-60 (µg)": 0.42,
    }

    with patch("grid_search.isotope_analysis.read_resnuclei_file", return_value=fake_row):
        run_isotope_analysis(tmp_path, config, state)

    # Find the "Normalized" title row, read the table immediately below it
    df_full = pd.read_excel(
        tmp_path / "isotopes.xlsx", sheet_name="Summary", header=None
    )
    norm_title_idx = df_full[
        df_full.iloc[:, 0].astype(str).str.contains("Normalized", na=False)
    ].index[0]
    df_norm = pd.read_excel(
        tmp_path / "isotopes.xlsx", sheet_name="Summary", header=norm_title_idx + 1
    )
    assert "Co-60 (Bq/cm³)" in df_norm.columns
    assert df_norm["Co-60 (Bq/cm³)"].iloc[0] == pytest.approx(200.0)  # 1000.0 / 5.0


def test_summary_sheet_absent_when_no_data(tmp_path):
    ia = IsotopeConfig(isotopes={27: 60}, rnc_files=["merged_21"])
    config = MagicMock()
    config.isotope_analysis = ia

    state = StateManager(tmp_path / "state.json")
    state.data = {"beame0.05_matGALLIUM": {"parameters": {}, "runs": {}}}

    with patch("grid_search.isotope_analysis.read_resnuclei_file", return_value=None):
        run_isotope_analysis(tmp_path, config, state)

    assert not (tmp_path / "isotopes.xlsx").exists()


def _make_summary_rows_simple():
    """2 isotopes, 2 cooling times, 2 beame values — no grouping param."""
    return [
        {"_tdecay_s": 0.0,     "CoolingTime": "0 s", "beame": 0.1, "Co-60 (Bq)": 100.0, "H-3 (Bq)": 10.0},
        {"_tdecay_s": 86400.0, "CoolingTime": "1 d", "beame": 0.1, "Co-60 (Bq)": 80.0,  "H-3 (Bq)": 8.0},
        {"_tdecay_s": 0.0,     "CoolingTime": "0 s", "beame": 0.5, "Co-60 (Bq)": 200.0, "H-3 (Bq)": 20.0},
        {"_tdecay_s": 86400.0, "CoolingTime": "1 d", "beame": 0.5, "Co-60 (Bq)": 160.0, "H-3 (Bq)": 16.0},
    ]


def test_build_pivot_sheet_creates_pivot_sheet(tmp_path):
    output = tmp_path / "pivot_test.xlsx"
    with pd.ExcelWriter(str(output), engine="openpyxl") as writer:
        _build_pivot_sheet(
            writer,
            summary_rows=_make_summary_rows_simple(),
            param_names=["beame"],
            group_by=None,
            volume=2.0,
        )

    wb = load_workbook(output)
    assert "Pivot" in wb.sheetnames


def test_build_pivot_sheet_bq_values_present(tmp_path):
    output = tmp_path / "pivot_test.xlsx"
    with pd.ExcelWriter(str(output), engine="openpyxl") as writer:
        _build_pivot_sheet(
            writer,
            summary_rows=_make_summary_rows_simple(),
            param_names=["beame"],
            group_by=None,
            volume=2.0,
        )

    wb = load_workbook(output)
    ws = wb["Pivot"]
    flat = [cell.value for row in ws.iter_rows() for cell in row]
    assert 100.0 in flat   # Co-60 Bq at beame=0.1, 0 s
    assert 200.0 in flat   # Co-60 Bq at beame=0.5, 0 s


def test_build_pivot_sheet_normalized_values_present(tmp_path):
    output = tmp_path / "pivot_test.xlsx"
    with pd.ExcelWriter(str(output), engine="openpyxl") as writer:
        _build_pivot_sheet(
            writer,
            summary_rows=_make_summary_rows_simple(),
            param_names=["beame"],
            group_by=None,
            volume=2.0,
        )

    wb = load_workbook(output)
    ws = wb["Pivot"]
    flat = [cell.value for row in ws.iter_rows() for cell in row]
    # 100.0 / 2.0 = 50.0 and 200.0 / 2.0 = 100.0
    assert any(abs(v - 50.0) < 1e-6 for v in flat if isinstance(v, float))
    assert any(abs(v - 100.0) < 1e-6 for v in flat if isinstance(v, float))


def test_build_pivot_sheet_title_row(tmp_path):
    output = tmp_path / "pivot_test.xlsx"
    with pd.ExcelWriter(str(output), engine="openpyxl") as writer:
        _build_pivot_sheet(
            writer,
            summary_rows=_make_summary_rows_simple(),
            param_names=["beame"],
            group_by=None,
            volume=2.0,
        )

    wb = load_workbook(output)
    ws = wb["Pivot"]
    assert ws.cell(row=1, column=1).value == "Activity (Bq)"


def test_build_pivot_sheet_cooling_time_order(tmp_path):
    """0 s must appear before 1 d in the sheet regardless of input order."""
    rows_reversed = [
        {"_tdecay_s": 86400.0, "CoolingTime": "1 d", "beame": 0.1, "Co-60 (Bq)": 80.0},
        {"_tdecay_s": 0.0,     "CoolingTime": "0 s", "beame": 0.1, "Co-60 (Bq)": 100.0},
    ]
    output = tmp_path / "pivot_order.xlsx"
    with pd.ExcelWriter(str(output), engine="openpyxl") as writer:
        _build_pivot_sheet(
            writer,
            summary_rows=rows_reversed,
            param_names=["beame"],
            group_by=None,
            volume=1.0,
        )

    wb = load_workbook(output)
    ws = wb["Pivot"]
    # CoolingTime values are in column 2 (index level 1)
    col2 = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    idx_zero = next((i for i, v in enumerate(col2) if v == "0 s"), None)
    idx_one  = next((i for i, v in enumerate(col2) if v == "1 d"), None)
    assert idx_zero is not None and idx_one is not None
    assert idx_zero < idx_one


def test_build_pivot_sheet_empty_rows_does_nothing(tmp_path):
    output = tmp_path / "pivot_empty.xlsx"
    with pd.ExcelWriter(str(output), engine="openpyxl") as writer:
        pd.DataFrame({"x": [1]}).to_excel(writer, sheet_name="dummy", index=False)
        _build_pivot_sheet(
            writer,
            summary_rows=[],
            param_names=["beame"],
            group_by=None,
            volume=1.0,
        )

    wb = load_workbook(output)
    assert "Pivot" not in wb.sheetnames


def test_build_pivot_sheet_group_by_titles(tmp_path):
    rows = [
        {"_tdecay_s": 0.0, "CoolingTime": "0 s", "mat": "GALLIUM",  "beame": 0.5, "Co-60 (Bq)": 100.0},
        {"_tdecay_s": 0.0, "CoolingTime": "0 s", "mat": "TUNGSTEN", "beame": 0.5, "Co-60 (Bq)": 50.0},
    ]
    output = tmp_path / "pivot_grouped.xlsx"
    with pd.ExcelWriter(str(output), engine="openpyxl") as writer:
        _build_pivot_sheet(
            writer,
            summary_rows=rows,
            param_names=["mat", "beame"],
            group_by="mat",
            volume=1.0,
        )

    wb = load_workbook(output)
    ws = wb["Pivot"]
    flat = [cell.value for row in ws.iter_rows() for cell in row]
    assert "mat=GALLIUM" in flat
    assert "mat=TUNGSTEN" in flat


def test_build_pivot_sheet_group_by_values(tmp_path):
    rows = [
        {"_tdecay_s": 0.0, "CoolingTime": "0 s", "mat": "GALLIUM",  "beame": 0.5, "Co-60 (Bq)": 100.0},
        {"_tdecay_s": 0.0, "CoolingTime": "0 s", "mat": "TUNGSTEN", "beame": 0.5, "Co-60 (Bq)": 50.0},
    ]
    output = tmp_path / "pivot_grouped_vals.xlsx"
    with pd.ExcelWriter(str(output), engine="openpyxl") as writer:
        _build_pivot_sheet(
            writer,
            summary_rows=rows,
            param_names=["mat", "beame"],
            group_by="mat",
            volume=1.0,
        )

    wb = load_workbook(output)
    ws = wb["Pivot"]
    flat = [cell.value for row in ws.iter_rows() for cell in row]
    assert 100.0 in flat
    assert 50.0 in flat
